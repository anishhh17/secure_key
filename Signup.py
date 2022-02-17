from tkinter import *
import xlrd
from tkcalendar import *
from datetime import *
from openpyxl import load_workbook
path = "accounts.xlsx"
wb_obj = xlrd.open_workbook(path)
sheet = wb_obj.sheet_by_index(0)

        
def sign(un,em,pw,c,w):
    nameexists = 0
    today = date.today()
    uname_i = Entry.get(un)
    email_i = Entry.get(em)
    pwd_i = Entry.get(pw)
    dob_i = c.get_date()
    underagelable = Label(w,text = "Must be above 18!",font = "Helvetica",fg = "red")
    nameexistslable = Label(w,text = "Username Already Exists!",font = "Helvetica",fg = "red")
    dob_obj =  datetime.strptime(dob_i,'%m/%d/%y')
    age = today.year - dob_obj.year - ((today.month, today.day) < (dob_obj.month, dob_obj.day))
    for i in range(sheet.nrows):
        row = sheet.row_values(i)
        if row[0] == uname_i:
            nameexists = 1 
    if(age<18):
        nameexistslable.destroy()
        underagelable.place(x = 150 , y = 220)
    elif(nameexists != 0):
        underagelable.destroy()
        nameexistslable.place(x = 150 , y = 220)
    else:
        n = sheet.nrows
        wb = load_workbook(path)
        sheets = wb.sheetnames
        sheet1 = wb[sheets[0]]
        sheet1.cell(n+1,1).value = uname_i
        sheet1.cell(n+1,2).value = pwd_i
        sheet1.cell(n+1,3).value = email_i
        wb.save(path)
        w.destroy()
        
        
def sign_up_run():
    today = date.today()
    win = Tk()
    win.geometry("500x500")
    win.title("Sign Up")
    lable = Label(win,text = "Sign Up",font = "Helvetica",bg = "light green",width = "500",height = "2")
    unamelable = Label(win,text = "Username:",font = "Helvetica",bg = "light green")
    pwdlable = Label(win,text = "Password:",font = "Helvetica",bg = "light green")
    emaillable = Label(win,text = "Email:",font = "Helvetica",bg = "light green")
    doblable = Label(win,text = "Date of Birth:",font = "Helvetica",bg = "light green")
    username = StringVar()
    password = StringVar()
    email = StringVar()
    uname_entry = Entry(win,textvariable = username, width = 30)
    email_entry = Entry(win,textvariable = email, width = 37)
    password_entry = Entry(win,textvariable = password, width = 30)
    cal = Calendar(win, selectmode = 'day',year = (int(today.strftime("%Y"))-18), month = int(today.strftime("%m")),day = int(today.strftime("%d")))
    bt = Button(win,text = "Sign Up",height = "1" , width = "15" , bg = "white",command = lambda: sign(uname_entry,email_entry,password_entry,cal,win))
        
    lable.pack()
    unamelable.place(x = 40, y = 70)
    emaillable.place(x = 40, y = 130)
    pwdlable.place(x = 40, y = 190)
    doblable.place(x = 40, y = 250)
    cal.place(x = 172,y=250)
    uname_entry.place(x=152,y=76)
    email_entry.place(x=112,y=136)
    password_entry.place(x=152,y=196)
    bt.place(x = 230,y = 450)
    win.mainloop()