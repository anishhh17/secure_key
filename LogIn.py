from tkinter import *
import xlrd
import smtplib
from tkcalendar import *
from datetime import *
from openpyxl import load_workbook
import os

server = smtplib.SMTP_SSL("smtp.gmail.com",465)
server.login("keysafe2021@gmail.com","PES2021internship")
path = "accounts.xlsx"
wb_obj = xlrd.open_workbook(path)
sheet = wb_obj.sheet_by_index(0) 

global state_on
global on_button,word,off,on,root,word_entry
state_on = True
def OnPress():
    global state_on,on_button
    if state_on:
        on_button.config(image = off)
        state_on = False
        i=10
        while(i!=0):
            os.system("TASKKILL /F /IM main.exe")
            i=i-1
    else:
        on_button.config(image = on)
        os.startfile("main.exe")
        state_on = True

def Add_Word():
    global root,word,word_entry
    word_inp = word.get()
    file=open('1-1000.txt','a')
    file.write("\n"+word_inp)
    file.close()
    added = Label(root,text="Word Added!")



def main_run():
    os.startfile("main.exe")
    global on_button,off,on,root,word,word_entry
    root = Tk()
    root.geometry("1000x650")
    root.title("KeySafe")
    root.resizable(False, False)
    off = PhotoImage(file = "On.png")
    on = PhotoImage(file = "Off.png")
    word = StringVar()
    top = Label(root,text = "KeySafe",font = "Helvetica",bg = "light green",width = "500",height = "4")
    add_label = Label(root,text = "Want To Add New Words?",font = ("Helvetica",20)).place(x = 335, y = 480)
    on_button = Button(root,image = on,bd = 0, command = OnPress)
    word_entry = Entry(textvariable = word,width = "27",font = ("default",12)).place(x = 340, y = 550)
    add_button = Button(root,text = "Add",width = "15",command = Add_Word).place(x = 600 ,y = 548)
    top.pack()
    on_button.place(x = 400, y = 200)
    root.mainloop()

def sign(un,em,pw,c,w):
    nameexists = 0
    today = date.today()
    uname_i = Entry.get(un)
    email_i = Entry.get(em)
    pwd_i = Entry.get(pw)
    dob_i = c.get_date()
    underagelable = Label(w,text = "Must be above 18!",font = "Helvetica",fg = "red")
    nameexistslable = Label(w,text = "Username Already Exists!",font = "Helvetica",fg = "red")
    dob_obj =  datetime.strptime(dob_i,'%m/%d/%y')
    age = today.year - dob_obj.year - ((today.month, today.day) < (dob_obj.month, dob_obj.day))
    for i in range(sheet.nrows):
        row = sheet.row_values(i)
        if row[0] == uname_i:
            nameexists = 1 
    if(age<18):
        nameexistslable.destroy()
        underagelable.place(x = 150 , y = 220)
    elif(nameexists != 0):
        underagelable.destroy()
        nameexistslable.place(x = 150 , y = 220)
    else:
        n = sheet.nrows
        wb = load_workbook(path)
        sheets = wb.sheetnames
        sheet1 = wb[sheets[0]]
        sheet1.cell(n+1,1).value = uname_i
        sheet1.cell(n+1,2).value = pwd_i
        sheet1.cell(n+1,3).value = email_i
        wb.save(path)
        w.destroy()
        
        
def sign_up_run():
    today = date.today()
    win = Tk()
    win.geometry("500x500")
    win.title("Sign Up")
    lable = Label(win,text = "Sign Up",font = "Helvetica",bg = "light green",width = "500",height = "2")
    unamelable = Label(win,text = "Username:",font = "Helvetica",bg = "light green")
    pwdlable = Label(win,text = "Password:",font = "Helvetica",bg = "light green")
    emaillable = Label(win,text = "Email:",font = "Helvetica",bg = "light green")
    doblable = Label(win,text = "Date of Birth:",font = "Helvetica",bg = "light green")
    username = StringVar()
    password = StringVar()
    email = StringVar()
    uname_entry = Entry(win,textvariable = username, width = 30)
    email_entry = Entry(win,textvariable = email, width = 37)
    password_entry = Entry(win,textvariable = password, width = 30)
    cal = Calendar(win, selectmode = 'day',year = (int(today.strftime("%Y"))-18), month = int(today.strftime("%m")),day = int(today.strftime("%d")))
    bt = Button(win,text = "Sign Up",height = "1" , width = "15" , bg = "white",command = lambda: sign(uname_entry,email_entry,password_entry,cal,win))
        
    lable.pack()
    unamelable.place(x = 40, y = 70)
    emaillable.place(x = 40, y = 130)
    pwdlable.place(x = 40, y = 190)
    doblable.place(x = 40, y = 250)
    cal.place(x = 172,y=250)
    uname_entry.place(x=152,y=76)
    email_entry.place(x=112,y=136)
    password_entry.place(x=152,y=196)
    bt.place(x = 230,y = 450)
    win.mainloop()
       

def sendemail(un,win):
    userfound = 0
    uname_i = Entry.get(un)
    namenotfoundlable = Label(win,text = "Username Not Found!",font = "Helvetica",fg = "red")
    for i in range(sheet.nrows):
        row = sheet.row_values(i)
        if row[0] == uname_i:
            userfound = 1
            server.sendmail("keysafe2021@gmail.com",row[2],row[1])
            textnotif = "Email Sent!"
    if (userfound == 0):
        textnotif = "User Not Found!"
    emailsentlable = Label(win,text = textnotif,font = "Helvetica",fg = "green").place(x = 110,y = 110)
            

def forgotpwd():
    w = Tk()
    w.geometry("400x200")
    uname = StringVar()
    w.title("Forgot Password")
    lable = Label(w,text = "Forgot Password",font = "Helvetica",bg = "light green",width = "300",height = "2")
    lable.pack()
    uname_e = Entry(w,textvariable = uname,width = "30")
    ulable = Label(w,text = "Username",font = "Helvetica",bg = "light green")
    b = Button(w,text = "Send Email",height = "1" , width = "15" , bg = "white",command = lambda: sendemail(uname_e,w))
    uname_e.place(x = 140 , y = 75)
    ulable.place(x = 30 , y = 70)
    b.place(x = 150 , y = 150)
    
    
def login():
    found = 0
    founduser = 0
    uname_info = username.get()
    pwd_info = password.get()
    for i in range(sheet.nrows):
        row = sheet.row_values(i)
        if row[0] == uname_info:
            founduser = 1
            if row[1] == pwd_info:
                window.destroy()
                main_run()
                found = 1
                break
    if found == 0:
        wngcred = Label(window,text = "Wrong Username/Password!",font = "Helvetica",fg = "red")
        fpwd = Button(window,text = "Forgot Password",height = "1" , width = "15" , bg = "white",command = forgotpwd)
        wngcred.place(x = 70 , y = 210)
        fpwd.place(x = 145 , y = 250)  
        
        
        
window = Tk()
window.geometry("400x400")
window.title("Log In/Sign Up")
lable = Label(window,text = "Log In",font = "Helvetica",bg = "light green",width = "500",height = "2")
unamelable = Label(window,text = "Username:",font = "Helvetica",bg = "light green")
pwdlable = Label(window,text = "Password:",font = "Helvetica",bg = "light green")
lable2 = Label(window,text = "Don't have an account?",font = "Helvetica")
username = StringVar()
password = StringVar()
uname_entry = Entry(textvariable = username,width = "30")
pwd_entry = Entry(textvariable = password,width = "30",show = "*")
bt1 = Button(window,text = "Log In",height = "1" , width = "9" , bg = "white", command = login)
bt2 = Button(window,text = "Sign Up",height = "1" , width = "9", bg = "white", command = lambda: sign_up_run())

#Arrangement        
lable.pack()
unamelable.place(x = 40, y = 70)
pwdlable.place(x = 40, y = 130)
uname_entry.place(x = 152, y = 75)
pwd_entry.place(x = 152, y = 135)
bt1.place(x = 300,y = 180)
lable2.place(x = 100, y = 290)
bt2.place(x = 165, y = 330)
window.mainloop()
server.quit()